"""
Template Generation Service

Generates Excel templates with pending/overdue assignments for bulk upload.
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Protection
from datetime import datetime
from typing import List, Dict, Optional
from io import BytesIO


class TemplateGenerationService:
    """Service for generating Excel templates for bulk data upload."""

    @staticmethod
    def generate_template(user, filter_type: str = 'pending') -> BytesIO:
        """
        Generate Excel template with assignments based on filter.

        Args:
            user: Current user object
            filter_type: 'overdue', 'pending', or 'overdue_and_pending'

        Returns:
            BytesIO: Excel file in memory
        """
        from ....models.data_assignment import DataPointAssignment
        from ....models.framework import FrameworkDataField
        from ....models.entity import Entity
        from ....models.dimension import FieldDimension

        # Get assignments based on filter
        assignments = TemplateGenerationService._get_assignments(user, filter_type)

        if not assignments:
            raise ValueError(f"No {filter_type} assignments found for this user")

        # Expand dimensional assignments
        from datetime import date
        from ....models.esg_data import ESGData

        today = date.today()
        rows = []

        for assignment in assignments:
            field = assignment.field
            entity = assignment.entity

            # Skip computed fields
            if field.is_computed:
                continue

            # Get valid reporting dates
            valid_dates = assignment.get_valid_reporting_dates()
            if not valid_dates:
                continue

            # Determine which dates to include based on filter type
            dates_to_include = []

            if filter_type == 'overdue':
                # Include only overdue dates without existing data
                overdue_dates = [d for d in valid_dates if d < today]
                for overdue_date in overdue_dates:
                    existing_data = ESGData.query.filter_by(
                        field_id=assignment.field_id,
                        entity_id=assignment.entity_id,
                        reporting_date=overdue_date,
                        is_draft=False
                    ).first()
                    if not existing_data:
                        dates_to_include.append(overdue_date)

            elif filter_type == 'pending':
                # Include next/nearest date if no data exists
                existing_data = ESGData.query.filter_by(
                    field_id=assignment.field_id,
                    entity_id=assignment.entity_id,
                    is_draft=False
                ).first()
                if not existing_data and valid_dates:
                    dates_to_include.append(valid_dates[0])

            else:  # overdue_and_pending
                # Include ALL overdue dates without data, plus next pending date
                overdue_dates = [d for d in valid_dates if d < today]
                pending_dates = [d for d in valid_dates if d >= today]

                # Add overdue dates without existing data
                for overdue_date in overdue_dates:
                    existing_data = ESGData.query.filter_by(
                        field_id=assignment.field_id,
                        entity_id=assignment.entity_id,
                        reporting_date=overdue_date,
                        is_draft=False
                    ).first()
                    if not existing_data:
                        dates_to_include.append(overdue_date)

                # Add next pending date if exists
                if pending_dates:
                    dates_to_include.append(pending_dates[0])

            # Create rows for each date
            for reporting_date in dates_to_include:
                # Check if field has dimensions
                field_dimensions = FieldDimension.query.filter_by(
                    field_id=field.field_id
                ).all()

                if field_dimensions:
                    # Expand dimensional combinations
                    dim_combinations = TemplateGenerationService._get_dimension_combinations(
                        field.field_id
                    )

                    for dim_combo in dim_combinations:
                        row = TemplateGenerationService._create_row(
                            field, entity, assignment, reporting_date, dim_combo
                        )
                        rows.append(row)
                else:
                    # Non-dimensional field
                    row = TemplateGenerationService._create_row(
                        field, entity, assignment, reporting_date, None
                    )
                    rows.append(row)

        # Check if any rows were generated after filtering
        if not rows:
            raise ValueError(
                f"No valid {filter_type} assignments found. "
                f"All assignments may be computed fields or have no valid reporting dates."
            )

        # Create Excel file
        return TemplateGenerationService._create_excel(rows, filter_type)

    @staticmethod
    def _get_assignments(user, filter_type: str):
        """Get assignments based on filter type."""
        from ....models.data_assignment import DataPointAssignment
        from ....models.esg_data import ESGData
        from datetime import date

        today = date.today()

        # Base query - active assignments for user's entity
        # Note: User has entity_id (singular), not entities (plural)
        base_query = DataPointAssignment.query.filter(
            DataPointAssignment.entity_id == user.entity_id,
            DataPointAssignment.series_status == 'active'
        )

        if filter_type == 'overdue':
            # Assignments with past due dates and no submitted data
            assignments = []
            for assignment in base_query.all():
                valid_dates = assignment.get_valid_reporting_dates()
                # Handle None return value (e.g., when assignment has no company)
                if valid_dates is None:
                    continue
                overdue_dates = [d for d in valid_dates if d < today]

                for overdue_date in overdue_dates:
                    # Check if data exists
                    existing_data = ESGData.query.filter_by(
                        field_id=assignment.field_id,
                        entity_id=assignment.entity_id,
                        reporting_date=overdue_date,
                        is_draft=False
                    ).first()

                    if not existing_data:
                        assignments.append(assignment)
                        break  # Only add assignment once

            return assignments

        elif filter_type == 'pending':
            # Assignments with no data submitted (not necessarily overdue)
            assignments = []
            for assignment in base_query.all():
                # Check if any data exists for this assignment
                existing_data = ESGData.query.filter_by(
                    field_id=assignment.field_id,
                    entity_id=assignment.entity_id,
                    is_draft=False
                ).first()

                if not existing_data:
                    assignments.append(assignment)

            return assignments

        else:  # overdue_and_pending
            return base_query.all()

    @staticmethod
    def _get_dimension_combinations(field_id: str) -> List[Dict]:
        """Get all dimension combinations for a field."""
        from ....models.dimension import FieldDimension, DimensionValue

        field_dimensions = FieldDimension.query.filter_by(
            field_id=field_id
        ).all()

        if not field_dimensions:
            return [{}]

        # Get all dimension values for each dimension
        dim_values_map = {}
        for fd in field_dimensions:
            dim_values = DimensionValue.query.filter_by(
                dimension_id=fd.dimension_id
            ).all()
            dim_values_map[fd.dimension.name] = [dv.value for dv in dim_values]

        # Generate cartesian product of all dimension combinations
        combinations = [{}]
        for dim_name, values in dim_values_map.items():
            new_combinations = []
            for combo in combinations:
                for value in values:
                    new_combo = combo.copy()
                    new_combo[dim_name] = value
                    new_combinations.append(new_combo)
            combinations = new_combinations

        return combinations

    @staticmethod
    def _create_row(field, entity, assignment, reporting_date, dimensions: Optional[Dict]) -> Dict:
        """Create a row dictionary for the template."""
        from datetime import date

        # Determine status based on reporting date
        today = date.today()
        status = 'OVERDUE' if reporting_date < today else 'PENDING'

        row = {
            'Field_Name': field.field_name,
            'Entity': entity.name,
            'Rep_Date': reporting_date.strftime('%Y-%m-%d'),
            'Value': '',
            'Unit': field.default_unit or '',
            'Notes': '',
            'Status': status,
            # Hidden columns
            'Field_ID': field.field_id,
            'Entity_ID': entity.id,
            'Assignment_ID': assignment.id
        }

        # Add dimension columns
        if dimensions:
            for dim_name, dim_value in dimensions.items():
                row[f'Dimension_{dim_name}'] = dim_value

        return row

    @staticmethod
    def _create_excel(rows: List[Dict], filter_type: str) -> BytesIO:
        """Create Excel file with data and instructions sheets."""

        # Create DataFrame
        df = pd.DataFrame(rows)

        # Verify DataFrame has data
        if df.empty:
            raise ValueError("Cannot create Excel template from empty data")

        # Reorder columns: visible columns first, then hidden
        visible_cols = ['Field_Name', 'Entity', 'Rep_Date']

        # Add dimension columns (only if they exist in DataFrame)
        dim_cols = [col for col in df.columns if col.startswith('Dimension_')]
        visible_cols.extend(sorted(dim_cols))

        # Add editable columns
        visible_cols.extend(['Value', 'Unit', 'Notes', 'Status'])

        # Hidden columns at the end
        hidden_cols = ['Field_ID', 'Entity_ID', 'Assignment_ID']

        # Only include columns that actually exist in the DataFrame
        all_cols = [col for col in (visible_cols + hidden_cols) if col in df.columns]
        df = df[all_cols]

        # Create Excel in memory
        output = BytesIO()

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Data Entry sheet
            df.to_excel(writer, sheet_name='Data Entry', index=False)

            # Instructions sheet
            instructions_df = TemplateGenerationService._create_instructions()
            instructions_df.to_excel(writer, sheet_name='Instructions', index=False, header=False)

        # Post-process with openpyxl for formatting
        output.seek(0)
        wb = load_workbook(output)
        ws = wb['Data Entry']

        # Apply formatting
        gray_fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
        header_font = Font(bold=True)

        # Format headers
        for cell in ws[1]:
            cell.font = header_font

        # Protect read-only columns (all except Value and Notes)
        editable_cols = ['Value', 'Notes']
        for col_idx, col_name in enumerate(df.columns, start=1):
            col_letter = ws.cell(row=1, column=col_idx).column_letter

            if col_name not in editable_cols:
                # Gray fill and protect
                for row in range(2, len(df) + 2):
                    cell = ws.cell(row=row, column=col_idx)
                    cell.fill = gray_fill
                    cell.protection = Protection(locked=True)

        # Hide ID columns
        for col_name in hidden_cols:
            col_idx = list(df.columns).index(col_name) + 1
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            ws.column_dimensions[col_letter].hidden = True

        # Enable sheet protection (allows editing unlocked cells)
        # Note: No password set - just UI-level protection
        ws.protection.sheet = True

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return output

    @staticmethod
    def _create_instructions() -> pd.DataFrame:
        """Create instructions sheet content."""
        instructions = [
            ["HOW TO USE THIS TEMPLATE"],
            [""],
            ["1. Fill in the 'Value' column (editable) for each row"],
            ["2. Optionally add notes in the 'Notes' column"],
            ["3. Do NOT modify other columns (marked in gray)"],
            ["4. Do NOT delete or add rows"],
            ["5. Save the file and upload it back to the dashboard"],
            [""],
            ["DIMENSIONAL DATA"],
            [""],
            ["• Some fields have dimension columns (e.g., Gender, Age)"],
            ["• One row = one dimension combination"],
            ["• Fill in value for EACH combination"],
            [""],
            ["VALIDATION RULES"],
            [""],
            ["• Values must match data type:"],
            ["  - INTEGER: Whole numbers only (e.g., 150)"],
            ["  - DECIMAL: Numbers with decimals (e.g., 1234.56)"],
            ["  - PERCENTAGE: Enter as 15 or 0.15 (both accepted)"],
            ["  - CURRENCY: Can include $ and commas (e.g., $1,000.50)"],
            ["  - BOOLEAN: TRUE/FALSE, YES/NO, or 1/0"],
            ["  - TEXT: Any text"],
            [""],
            ["• All required fields must have values"],
            ["• Dates cannot be modified"],
            ["• Notes limited to 1000 characters"],
            [""],
            ["AFTER UPLOAD"],
            [""],
            ["• You'll be able to attach files for each field"],
            ["• Same file can be uploaded multiple times if needed"],
            ["• All changes will be validated before saving"],
        ]

        return pd.DataFrame(instructions)
