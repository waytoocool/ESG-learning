#!/usr/bin/env python3
"""
Test 1: SQL Injection in Notes Field
TC-EH-007
"""

import pandas as pd
import openpyxl
from datetime import datetime

# Read the downloaded template
template_path = '.playwright-mcp/Template-overdue-2025-11-19.xlsx'
wb = openpyxl.load_workbook(template_path)
ws = wb.active

# Print headers to understand structure
print("Template headers:")
for idx, cell in enumerate(ws[1], 1):
    print(f"  Column {idx}: {cell.value}")

# Find the first data row (should be row 2)
print("\nOriginal row 2 data:")
for idx, cell in enumerate(ws[2], 1):
    print(f"  Column {idx}: {cell.value}")

# Locate the Value and Notes columns
header_row = [cell.value for cell in ws[1]]
value_col_idx = None
notes_col_idx = None

for idx, header in enumerate(header_row, 1):
    if header and 'value' in str(header).lower():
        value_col_idx = idx
    if header and 'notes' in str(header).lower():
        notes_col_idx = idx

print(f"\nValue column: {value_col_idx}")
print(f"Notes column: {notes_col_idx}")

# Inject SQL injection payload
if value_col_idx and notes_col_idx:
    # Set value
    ws.cell(row=2, column=value_col_idx).value = 100

    # Set SQL injection payload in notes
    sql_injection = "'; DROP TABLE esg_data; --"
    ws.cell(row=2, column=notes_col_idx).value = sql_injection

    print(f"\nInjected SQL payload in row 2:")
    print(f"  Value: {ws.cell(row=2, column=value_col_idx).value}")
    print(f"  Notes: {ws.cell(row=2, column=notes_col_idx).value}")

    # Save the modified file
    output_path = '.playwright-mcp/Template-overdue-SQL-INJECTION.xlsx'
    wb.save(output_path)
    print(f"\nSaved modified template to: {output_path}")
else:
    print("\nERROR: Could not find Value or Notes columns!")
