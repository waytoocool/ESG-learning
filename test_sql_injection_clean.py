#!/usr/bin/env python3
"""
Test 1: SQL Injection in Notes Field - Clean version with only one row
TC-EH-007
"""

import openpyxl

# Read the downloaded template
template_path = '.playwright-mcp/Template-overdue-2025-11-19.xlsx'
wb = openpyxl.load_workbook(template_path)
ws = wb.active

# Find the Value and Notes columns
header_row = [cell.value for cell in ws[1]]
value_col_idx = None
notes_col_idx = None

for idx, header in enumerate(header_row, 1):
    if header and 'value' in str(header).lower():
        value_col_idx = idx
    if header and 'notes' in str(header).lower():
        notes_col_idx = idx

print(f"Value column: {value_col_idx}")
print(f"Notes column: {notes_col_idx}")

# Delete all rows except header and row 2
max_row = ws.max_row
for row_idx in range(max_row, 2, -1):  # Delete from bottom to top
    ws.delete_rows(row_idx)

print(f"\nRows after deletion: {ws.max_row}")

# Inject SQL injection payload in row 2
if value_col_idx and notes_col_idx:
    # Set value
    ws.cell(row=2, column=value_col_idx).value = 100

    # Set SQL injection payload in notes
    sql_injection = "'; DROP TABLE esg_data; --"
    ws.cell(row=2, column=notes_col_idx).value = sql_injection

    print(f"\nInjected SQL payload in row 2:")
    print(f"  Value: {ws.cell(row=2, column=value_col_idx).value}")
    print(f"  Notes: {ws.cell(row=2, column=notes_col_idx).value}")

    # Print all data in row 2
    print("\nComplete row 2 data:")
    for idx, cell in enumerate(ws[2], 1):
        if cell.value:
            print(f"  Column {idx} ({ws.cell(row=1, column=idx).value}): {cell.value}")

    # Save the modified file
    output_path = '.playwright-mcp/Template-SQL-INJECTION-TEST.xlsx'
    wb.save(output_path)
    print(f"\nSaved modified template to: {output_path}")
else:
    print("\nERROR: Could not find Value or Notes columns!")
