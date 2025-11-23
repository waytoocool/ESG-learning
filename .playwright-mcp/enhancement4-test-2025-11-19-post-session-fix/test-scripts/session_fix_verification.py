#!/usr/bin/env python3
"""
PRIORITY 1: Session Fix Verification Test
Tests the critical session persistence bug fix in bulk_upload_api.py line 245
"""

import requests
import openpyxl
import time
import sqlite3
import os
from datetime import datetime

# Configuration
BASE_URL = "http://test-company-alpha.127-0-0-1.nip.io:8000"
LOGIN_EMAIL = "bob@alpha.com"
LOGIN_PASSWORD = "user123"
DB_PATH = "instance/esg_data.db"

# Test directories
TEST_DIR = "/Users/prateekgoyal/Desktop/Prateek/ESG DataVault Development/Claude/sakshi-learning/.playwright-mcp/enhancement4-test-2025-11-19-post-session-fix"
TEMPLATES_DIR = f"{TEST_DIR}/templates"
SCREENSHOTS_DIR = f"{TEST_DIR}/screenshots"
DB_VERIFICATION_DIR = f"{TEST_DIR}/database-verification"

# Create session
session = requests.Session()

def log(message):
    """Print timestamped log message"""
    timestamp = datetime.now().strftime("%H:%M:%S")
    print(f"[{timestamp}] {message}")

def login():
    """Login as test user"""
    log("Step 1: Logging in as bob@alpha.com...")

    # Get login page to get CSRF token
    response = session.get(f"{BASE_URL}/login")
    if response.status_code != 200:
        log(f"❌ Failed to access login page: {response.status_code}")
        return False

    # Login
    login_data = {
        "email": LOGIN_EMAIL,
        "password": LOGIN_PASSWORD
    }

    response = session.post(f"{BASE_URL}/login", data=login_data, allow_redirects=True)

    if "dashboard" in response.url or response.status_code == 200:
        log("✅ Login successful")
        return True
    else:
        log(f"❌ Login failed: {response.status_code}")
        log(f"Response URL: {response.url}")
        return False

def download_template():
    """Download pending only template"""
    log("\nStep 2: Downloading 'Pending Only' template...")

    # Use POST request with JSON body
    response = session.post(
        f"{BASE_URL}/api/user/v2/bulk-upload/template",
        json={"filter": "pending"},
        headers={'Content-Type': 'application/json'}
    )

    if response.status_code == 200:
        timestamp = str(int(time.time()))
        template_path = f"{TEMPLATES_DIR}/Template-pending-{timestamp}.xlsx"

        with open(template_path, 'wb') as f:
            f.write(response.content)

        log(f"✅ Template downloaded: {template_path}")
        return template_path
    else:
        log(f"❌ Template download failed: {response.status_code}")
        return None

def fill_template(template_path):
    """Fill template with test data"""
    log("\nStep 3: Filling template with test data...")

    try:
        wb = openpyxl.load_workbook(template_path)
        ws = wb['Data Entry']

        timestamp = str(int(time.time()))
        rows_filled = 0

        # Fill first 3 data rows
        for idx in range(2, 5):  # Rows 2-4 (row 1 is header)
            # Check if row has data
            if ws.cell(row=idx, column=1).value:  # If Field_Name exists
                ws.cell(row=idx, column=4).value = 100 + idx  # Value column
                ws.cell(row=idx, column=6).value = f'SESSION-FIX-TEST-{timestamp}-ROW-{idx}'  # Notes column
                rows_filled += 1

        if rows_filled == 0:
            log("❌ No rows found in template to fill")
            return None

        filled_template_path = f"{TEMPLATES_DIR}/Template-pending-SESSIONFIX-{timestamp}.xlsx"
        wb.save(filled_template_path)

        log(f"✅ Template filled with {rows_filled} rows of test data")
        log(f"   Saved to: {filled_template_path}")
        return filled_template_path, timestamp

    except Exception as e:
        log(f"❌ Error filling template: {str(e)}")
        return None

def upload_template(file_path):
    """Upload filled template"""
    log("\nStep 4: Uploading filled template...")

    try:
        with open(file_path, 'rb') as f:
            files = {'file': (os.path.basename(file_path), f, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
            response = session.post(f"{BASE_URL}/api/user/v2/bulk-upload/upload", files=files)

        if response.status_code == 200:
            data = response.json()
            log(f"✅ Upload successful")
            log(f"   Upload ID: {data.get('upload_id', 'N/A')}")
            log(f"   Total rows: {data.get('total_rows', 0)}")
            return data.get('upload_id')
        else:
            log(f"❌ Upload failed: {response.status_code}")
            log(f"   Response: {response.text}")
            return None

    except Exception as e:
        log(f"❌ Error uploading template: {str(e)}")
        return None

def validate_data(upload_id):
    """Validate uploaded data"""
    log("\nStep 5: Validating uploaded data...")

    try:
        response = session.post(
            f"{BASE_URL}/api/user/v2/bulk-upload/validate",
            json={"upload_id": upload_id},
            headers={'Content-Type': 'application/json'}
        )

        if response.status_code == 200:
            data = response.json()
            log(f"✅ Validation successful")
            log(f"   Valid rows: {data.get('valid_count', 0)}")
            log(f"   Invalid rows: {data.get('invalid_count', 0)}")
            log(f"   Warning rows: {data.get('warning_count', 0)}")
            log(f"   Overwrite rows: {data.get('overwrite_count', 0)}")

            if data.get('valid_count', 0) > 0:
                return True
            else:
                log("❌ No valid rows found after validation")
                return False
        else:
            log(f"❌ Validation failed: {response.status_code}")
            log(f"   Response: {response.text}")
            return False

    except Exception as e:
        log(f"❌ Error validating data: {str(e)}")
        return False

def submit_data(upload_id):
    """Submit validated data - THIS IS THE CRITICAL TEST"""
    log("\nStep 6: Submitting data (CRITICAL SESSION FIX TEST)...")

    try:
        # Submit with upload_id as form data
        response = session.post(
            f"{BASE_URL}/api/user/v2/bulk-upload/submit",
            data={"upload_id": upload_id}
        )

        log(f"   Response status: {response.status_code}")

        if response.status_code == 200:
            data = response.json()
            log(f"   Response data: {data}")

            # Check for the specific error that indicates session bug
            if data.get('success') == False and 'No validated rows found' in data.get('message', ''):
                log("❌ CRITICAL: 'No validated rows found' error - SESSION BUG STILL PRESENT")
                return False

            # Check for success
            if data.get('success') == True:
                log(f"✅ SUBMISSION SUCCESSFUL!")
                log(f"   Created: {data.get('created_count', 0)}")
                log(f"   Updated: {data.get('updated_count', 0)}")
                log(f"   Batch ID: {data.get('batch_id', 'N/A')}")
                return True
            else:
                log(f"❌ Submission failed: {data.get('message', 'Unknown error')}")
                return False
        else:
            log(f"❌ Submission failed with status {response.status_code}")
            log(f"   Response: {response.text}")
            return False

    except Exception as e:
        log(f"❌ Error submitting data: {str(e)}")
        return False

def verify_database(timestamp):
    """Verify data in database"""
    log("\nStep 7: DATABASE VERIFICATION...")

    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()

        # Find entries with our test timestamp
        cursor.execute("""
            SELECT data_id, raw_value, notes, created_at
            FROM esg_data
            WHERE notes LIKE ?
            ORDER BY created_at DESC
        """, (f'%SESSION-FIX-TEST-{timestamp}%',))

        results = cursor.fetchall()

        log(f"\n{'='*60}")
        log(f"DATABASE VERIFICATION RESULTS")
        log(f"{'='*60}")
        log(f"Entries found: {len(results)}")

        if len(results) > 0:
            log(f"\nEntries:")
            for row in results:
                log(f"  ✓ data_id={row[0]}")
                log(f"    value={row[1]}")
                log(f"    notes={row[2][:60]}...")
                log(f"    created={row[3]}")

        log(f"{'='*60}\n")

        # Save results to file
        results_file = f"{DB_VERIFICATION_DIR}/session_fix_verification_{timestamp}.txt"
        with open(results_file, 'w') as f:
            f.write(f"DATABASE VERIFICATION RESULTS\n")
            f.write(f"{'='*60}\n")
            f.write(f"Test timestamp: {timestamp}\n")
            f.write(f"Entries found: {len(results)}\n\n")

            for row in results:
                f.write(f"data_id: {row[0]}\n")
                f.write(f"raw_value: {row[1]}\n")
                f.write(f"notes: {row[2]}\n")
                f.write(f"created_at: {row[3]}\n")
                f.write(f"-" * 60 + "\n")

        log(f"Database results saved to: {results_file}")

        conn.close()

        # Check if we got expected number of entries (3)
        if len(results) == 3:
            log("✅ SESSION FIX VERIFIED - 3 entries found in database!")
            return True
        else:
            log(f"❌ SESSION FIX FAILED - Expected 3 entries, found {len(results)}")
            return False

    except Exception as e:
        log(f"❌ Error verifying database: {str(e)}")
        return False

def main():
    """Execute complete session fix verification test"""
    log("="*60)
    log("PRIORITY 1: SESSION FIX VERIFICATION TEST")
    log("="*60)
    log(f"Testing critical bug fix in bulk_upload_api.py line 245")
    log(f"Expected: session.modified = True enables data submission")
    log("="*60 + "\n")

    # Step 1: Login
    if not login():
        log("\n❌ TEST FAILED: Cannot proceed without login")
        return False

    # Step 2: Download template
    template_path = download_template()
    if not template_path:
        log("\n❌ TEST FAILED: Cannot download template")
        return False

    # Step 3: Fill template
    result = fill_template(template_path)
    if not result:
        log("\n❌ TEST FAILED: Cannot fill template")
        return False

    filled_template_path, timestamp = result

    # Step 4: Upload template
    upload_id = upload_template(filled_template_path)
    if not upload_id:
        log("\n❌ TEST FAILED: Cannot upload template")
        return False

    # Step 5: Validate data
    if not validate_data(upload_id):
        log("\n❌ TEST FAILED: Validation failed")
        return False

    # Step 6: Submit data (CRITICAL TEST)
    submission_success = submit_data(upload_id)

    # Step 7: Verify database
    database_verified = verify_database(timestamp)

    # Final verdict
    log("\n" + "="*60)
    log("FINAL TEST RESULTS")
    log("="*60)

    if submission_success and database_verified:
        log("✅✅✅ SESSION FIX VERIFIED - ALL CHECKS PASSED ✅✅✅")
        log("   - Submission succeeded without 'No validated rows' error")
        log("   - 3 entries found in database")
        log("   - Session persistence fix is working correctly")
        log("\n🎉 READY TO PROCEED TO COMPREHENSIVE TESTING")
        return True
    else:
        log("❌❌❌ SESSION FIX FAILED ❌❌❌")
        if not submission_success:
            log("   - Submission failed or returned error")
        if not database_verified:
            log("   - Database verification failed (wrong number of entries)")
        log("\n🛑 STOP - DO NOT PROCEED TO COMPREHENSIVE TESTING")
        return False

if __name__ == "__main__":
    success = main()
    exit(0 if success else 1)
