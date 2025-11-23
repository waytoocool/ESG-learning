#!/usr/bin/env python3
"""
Test script to verify BUG-ENH4-006: Combined Template Status Fix
Inspects both Overdue-only and Overdue+Pending templates
"""
import openpyxl
from datetime import datetime
from collections import Counter

def inspect_template(file_path, template_name):
    print("\n" + "="*80)
    print(f"{template_name} INSPECTION - BUG-ENH4-006 VERIFICATION")
    print("="*80)

    wb = openpyxl.load_workbook(file_path)
    ws = wb['Data Entry']

    # Find Status column
    header_row = []
    status_col_idx = None
    date_col_idx = None

    for col_idx in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=1, column=col_idx).value
        header_row.append(cell_value)
        if cell_value and 'status' in str(cell_value).lower():
            status_col_idx = col_idx
        if cell_value and 'rep_date' in str(cell_value).lower():
            date_col_idx = col_idx

    print(f"\n✓ Status column at index: {status_col_idx}")
    print(f"✓ Rep_Date column at index: {date_col_idx}")
    print(f"✓ Total rows in template: {ws.max_row}")
    print(f"✓ Data rows: {ws.max_row - 1}")

    # Collect all status values
    status_values = []
    status_date_pairs = []

    for row_idx in range(2, ws.max_row + 1):
        status = ws.cell(row=row_idx, column=status_col_idx).value
        rep_date = ws.cell(row=row_idx, column=date_col_idx).value
        field_name = ws.cell(row=row_idx, column=1).value

        status_values.append(status)
        status_date_pairs.append({
            'row': row_idx,
            'status': status,
            'date': rep_date,
            'field': field_name
        })

    # Count status distribution
    status_counts = Counter(status_values)

    print(f"\n📊 Status Distribution:")
    for status, count in status_counts.items():
        print(f"  - {status}: {count} rows")

    # Show first 10 rows for verification
    print(f"\n📋 First 10 rows (Status | Date | Field):")
    print("-" * 80)
    for i, pair in enumerate(status_date_pairs[:10], 1):
        print(f"  Row {pair['row']}: {pair['status']:10} | {str(pair['date'])[:10]} | {pair['field'][:50]}")

    # Verify logic
    print(f"\n🔍 Verification:")
    today = datetime.now().date()
    errors = []

    for pair in status_date_pairs:
        status = pair['status']
        rep_date = pair['date']

        # Parse date
        if isinstance(rep_date, datetime):
            date_obj = rep_date.date()
        elif isinstance(rep_date, str):
            try:
                date_obj = datetime.fromisoformat(rep_date.replace('Z', '+00:00')).date()
            except:
                date_obj = None
        else:
            date_obj = None

        if date_obj:
            expected_status = 'OVERDUE' if date_obj < today else 'PENDING'
            if status != expected_status:
                errors.append({
                    'row': pair['row'],
                    'date': rep_date,
                    'expected': expected_status,
                    'actual': status
                })

    if errors:
        print(f"  ❌ FAILED - Found {len(errors)} rows with incorrect status:")
        for error in errors[:5]:  # Show first 5 errors
            print(f"    Row {error['row']}: Date={error['date']}, Expected={error['expected']}, Got={error['actual']}")
        return False
    else:
        print(f"  ✅ PASSED - All {len(status_date_pairs)} rows have correct status based on date")
        return True

# Test both templates
overdue_path = "/Users/prateekgoyal/Desktop/Prateek/ESG DataVault Development/Claude/sakshi-learning/.playwright-mcp/enhancement4-test-2025-11-19-final/templates-downloaded/Template-overdue-2025-11-19.xlsx"
combined_path = "/Users/prateekgoyal/Desktop/Prateek/ESG DataVault Development/Claude/sakshi-learning/.playwright-mcp/enhancement4-test-2025-11-19-final/templates-downloaded/Template-overdue-and-pending-2025-11-19.xlsx"

print("\n" + "#"*80)
print("# BUG-ENH4-006 VERIFICATION: Template Status Column Fix")
print("#"*80)

result1 = inspect_template(overdue_path, "OVERDUE ONLY TEMPLATE")
result2 = inspect_template(combined_path, "OVERDUE + PENDING TEMPLATE")

print("\n" + "="*80)
print("FINAL RESULTS - BUG-ENH4-006")
print("="*80)
print(f"Overdue Only Template: {'✅ PASS' if result1 else '❌ FAIL'}")
print(f"Overdue + Pending Template: {'✅ PASS' if result2 else '❌ FAIL'}")

if result1 and result2:
    print("\n🎉 BUG-ENH4-006 VERIFICATION: PASSED")
    print("All templates show correct status values based on reporting dates")
else:
    print("\n❌ BUG-ENH4-006 VERIFICATION: FAILED")
    print("Some templates have incorrect status values")
