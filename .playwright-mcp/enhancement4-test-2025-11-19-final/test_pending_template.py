#!/usr/bin/env python3
"""
Test script to:
1. Inspect Pending template status column (BUG-ENH4-004)
2. Fill template with test data for upload (BUG-ENH4-005)
"""
import openpyxl
from datetime import datetime

# Load the pending template
template_path = "/Users/prateekgoyal/Desktop/Prateek/ESG DataVault Development/Claude/sakshi-learning/.playwright-mcp/enhancement4-test-2025-11-19-final/templates-downloaded/Template-pending-2025-11-19.xlsx"
output_path = "/Users/prateekgoyal/Desktop/Prateek/ESG DataVault Development/Claude/sakshi-learning/.playwright-mcp/enhancement4-test-2025-11-19-final/templates-filled/Template-pending-FILLED.xlsx"

print("Loading template...")
wb = openpyxl.load_workbook(template_path)
ws = wb['Data Entry']

print("\n" + "="*80)
print("PENDING TEMPLATE INSPECTION - BUG-ENH4-004 VERIFICATION")
print("="*80)

# Print header row
print("\nHeader Row:")
header_row = []
for col_idx in range(1, ws.max_column + 1):
    cell_value = ws.cell(row=1, column=col_idx).value
    header_row.append(cell_value)
    print(f"  Col {col_idx}: {cell_value}")

# Find Status column
status_col_idx = None
for idx, header in enumerate(header_row, start=1):
    if header and 'status' in str(header).lower():
        status_col_idx = idx
        print(f"\n✓ Found Status column at index: {status_col_idx}")
        break

if not status_col_idx:
    print("\n✗ ERROR: Status column not found!")
else:
    # Inspect first 5 data rows
    print(f"\nFirst 5 rows Status values:")
    print("-" * 80)
    for row_idx in range(2, min(7, ws.max_row + 1)):
        status_value = ws.cell(row=row_idx, column=status_col_idx).value
        field_name = ws.cell(row=row_idx, column=1).value  # Assuming field name in col 1
        reporting_date = ws.cell(row=row_idx, column=3).value  # Assuming reporting date in col 3
        print(f"Row {row_idx}: Status='{status_value}' | Field='{field_name}' | Date={reporting_date}")

    # Check if all are PENDING (as per bug fix requirement)
    all_pending = True
    for row_idx in range(2, ws.max_row + 1):
        status_value = ws.cell(row=row_idx, column=status_col_idx).value
        if status_value and status_value != 'PENDING':
            all_pending = False
            print(f"\n⚠️  WARNING: Row {row_idx} has status '{status_value}' instead of 'PENDING'")

    if all_pending:
        print("\n✅ BUG-ENH4-004 VERIFICATION: PASS - All rows show 'PENDING' status")
    else:
        print("\n❌ BUG-ENH4-004 VERIFICATION: FAIL - Not all rows are PENDING")

print("\n" + "="*80)
print("FILLING TEMPLATE WITH TEST DATA - BUG-ENH4-005 TEST PREPARATION")
print("="*80)

# Fill first 3 data rows with test values
# Assuming: Col 1=Field Name, Col 2=Reporting Period, Col 3=Reporting Date, Col 4=Value, Col 5=Units, Col 6=Notes
print("\nFilling rows with test data...")
for row_idx in range(2, 5):  # Rows 2-4
    # Find Value column (usually col 4)
    value_col = 4
    notes_col = 6

    # Set value
    test_value = 100 + row_idx
    ws.cell(row=row_idx, column=value_col).value = test_value

    # Set notes
    ws.cell(row=row_idx, column=notes_col).value = f'Test data for row {row_idx} - automated test'

    field_name = ws.cell(row=row_idx, column=1).value
    print(f"  Row {row_idx}: Set value={test_value}, notes added | Field: {field_name}")

# Save filled template
print(f"\nSaving filled template to: {output_path}")
wb.save(output_path)
print("✅ Template filled and saved successfully!")

print("\n" + "="*80)
print("SUMMARY")
print("="*80)
print(f"Total rows in template: {ws.max_row}")
print(f"Filled rows: 2-4 (3 rows)")
print(f"Ready for upload test (BUG-ENH4-005)")
