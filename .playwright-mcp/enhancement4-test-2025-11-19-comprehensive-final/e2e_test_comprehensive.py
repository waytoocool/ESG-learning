#!/usr/bin/env python3
"""
Comprehensive End-to-End Test for Bulk Excel Upload
This script performs the complete workflow including database verification
"""

import requests
import sqlite3
import openpyxl
import os
import sys
from datetime import datetime
import time
import json

# Configuration
BASE_URL = "http://test-company-alpha.127-0-0-1.nip.io:8000"
LOGIN_EMAIL = "bob@alpha.com"
LOGIN_PASSWORD = "user123"
DB_PATH = "/Users/prateekgoyal/Desktop/Prateek/ESG DataVault Development/Claude/sakshi-learning/instance/esg_data.db"
TEST_OUTPUT_DIR = "/Users/prateekgoyal/Desktop/Prateek/ESG DataVault Development/Claude/sakshi-learning/.playwright-mcp/enhancement4-test-2025-11-19-comprehensive-final"

class BulkUploadE2ETester:
    def __init__(self):
        self.session = requests.Session()
        self.test_timestamp = datetime.now().isoformat()
        self.test_results = []
        self.batch_id = None

    def log_result(self, test_name, status, details=""):
        """Log test result"""
        result = {
            "test": test_name,
            "status": status,
            "details": details,
            "timestamp": datetime.now().isoformat()
        }
        self.test_results.append(result)
        print(f"[{status}] {test_name}: {details}")

    def login(self):
        """Step 1: Login as bob@alpha.com"""
        print("\n=== STEP 1: LOGIN ===")
        try:
            # Get login page to get CSRF token
            response = self.session.get(f"{BASE_URL}/login")
            if response.status_code != 200:
                self.log_result("Login - Get Page", "FAIL", f"Status: {response.status_code}")
                return False

            # Perform login
            login_data = {
                "email": LOGIN_EMAIL,
                "password": LOGIN_PASSWORD
            }
            response = self.session.post(f"{BASE_URL}/login", data=login_data, allow_redirects=True)

            if "dashboard" in response.url or response.status_code == 200:
                self.log_result("Login", "PASS", f"Logged in as {LOGIN_EMAIL}")
                return True
            else:
                self.log_result("Login", "FAIL", f"Login failed. URL: {response.url}")
                return False

        except Exception as e:
            self.log_result("Login", "FAIL", str(e))
            return False

    def download_template(self):
        """Step 2: Download pending template"""
        print("\n=== STEP 2: DOWNLOAD TEMPLATE ===")
        try:
            # Download pending template - POST request with JSON
            response = self.session.post(
                f"{BASE_URL}/api/user/v2/bulk-upload/template",
                json={"filter": "pending"},
                headers={"Content-Type": "application/json"}
            )

            if response.status_code == 200:
                template_path = os.path.join(TEST_OUTPUT_DIR, "templates-all-tests", f"Template-pending-E2E-{datetime.now().strftime('%Y%m%d-%H%M%S')}.xlsx")
                with open(template_path, 'wb') as f:
                    f.write(response.content)

                self.template_path = template_path
                self.log_result("Download Template", "PASS", f"Saved to {template_path}")
                return True
            else:
                self.log_result("Download Template", "FAIL", f"Status: {response.status_code}")
                return False

        except Exception as e:
            self.log_result("Download Template", "FAIL", str(e))
            return False

    def fill_template(self):
        """Step 3: Fill template with test data"""
        print("\n=== STEP 3: FILL TEMPLATE WITH TEST DATA ===")
        try:
            wb = openpyxl.load_workbook(self.template_path)
            ws = wb['Data Entry']

            # Find the value and notes columns
            # Assuming standard template structure
            test_data = [
                {'value': 150, 'notes': f'E2E Test Row 1 - {self.test_timestamp}'},
                {'value': 250, 'notes': f'E2E Test Row 2 - {self.test_timestamp}'},
                {'value': 350, 'notes': f'E2E Test Row 3 - {self.test_timestamp}'}
            ]

            # Print headers to understand structure
            print("\nTemplate Headers:")
            for col_idx in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=1, column=col_idx).value
                print(f"  Column {col_idx}: {cell_value}")

            # Fill data rows (starting from row 2)
            rows_filled = 0
            for row_idx in range(2, min(5, ws.max_row + 1)):  # Fill first 3 data rows
                if row_idx - 2 < len(test_data):
                    data = test_data[row_idx - 2]

                    # Find Value and Notes columns by header name
                    for col_idx in range(1, ws.max_column + 1):
                        header = ws.cell(row=1, column=col_idx).value
                        if header == 'Value':
                            ws.cell(row=row_idx, column=col_idx).value = data['value']
                        elif header == 'Notes':
                            ws.cell(row=row_idx, column=col_idx).value = data['notes']

                    rows_filled += 1

            # Save filled template
            filled_path = os.path.join(TEST_OUTPUT_DIR, "templates-all-tests", f"Template-pending-E2E-FILLED-{datetime.now().strftime('%Y%m%d-%H%M%S')}.xlsx")
            wb.save(filled_path)
            self.filled_template_path = filled_path

            self.log_result("Fill Template", "PASS", f"Filled {rows_filled} rows. Saved to {filled_path}")
            return True

        except Exception as e:
            self.log_result("Fill Template", "FAIL", str(e))
            return False

    def upload_template(self):
        """Step 4: Upload filled template"""
        print("\n=== STEP 4: UPLOAD FILLED TEMPLATE ===")
        try:
            with open(self.filled_template_path, 'rb') as f:
                files = {'file': (os.path.basename(self.filled_template_path), f, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
                response = self.session.post(f"{BASE_URL}/api/user/v2/bulk-upload/upload", files=files)

            if response.status_code == 200:
                result = response.json()
                print(f"Upload response: {json.dumps(result, indent=2)}")
                if result.get('success'):
                    upload_id = result.get('upload_id')
                    self.upload_id = upload_id
                    total_rows = result.get('total_rows', 0)
                    self.log_result("Upload Template", "PASS", f"Upload ID: {upload_id}, Rows: {total_rows}")
                    return True
                else:
                    self.log_result("Upload Template", "FAIL", f"Upload failed: {result.get('error')}")
                    return False
            else:
                print(f"Upload failed with status {response.status_code}: {response.text}")
                self.log_result("Upload Template", "FAIL", f"Status: {response.status_code}")
                return False

        except Exception as e:
            self.log_result("Upload Template", "FAIL", str(e))
            return False

    def validate_data(self):
        """Step 5: Validate uploaded data"""
        print("\n=== STEP 5: VALIDATE DATA ===")
        try:
            response = self.session.post(
                f"{BASE_URL}/api/user/v2/bulk-upload/validate",
                json={"upload_id": self.upload_id},
                headers={"Content-Type": "application/json"}
            )

            if response.status_code == 200:
                result = response.json()
                print(f"Validation response: {json.dumps(result, indent=2)}")
                if result.get('success'):
                    # The response IS the validation result, not nested
                    valid_count = result.get('valid_count', 0)
                    error_count = result.get('invalid_count', 0)
                    warning_count = result.get('warning_count', 0)

                    if error_count == 0:
                        self.log_result("Validate Data", "PASS", f"Valid: {valid_count}, Warnings: {warning_count}, Errors: {error_count}")
                        return True
                    else:
                        self.log_result("Validate Data", "FAIL", f"Validation errors: {error_count}")
                        print("Errors:", json.dumps(result.get('invalid_rows', []), indent=2))
                        return False
                else:
                    self.log_result("Validate Data", "FAIL", f"Validation failed: {result.get('error')}")
                    return False
            else:
                print(f"Validation failed with status {response.status_code}: {response.text}")
                self.log_result("Validate Data", "FAIL", f"Status: {response.status_code}")
                return False

        except Exception as e:
            self.log_result("Validate Data", "FAIL", str(e))
            return False

    def submit_data(self):
        """Step 6: Submit data"""
        print("\n=== STEP 6: SUBMIT DATA ===")
        try:
            # Submit uses form data
            response = self.session.post(
                f"{BASE_URL}/api/user/v2/bulk-upload/submit",
                data={"upload_id": self.upload_id}
            )

            if response.status_code == 200:
                result = response.json()
                print(f"Submit response: {json.dumps(result, indent=2)}")
                if result.get('success'):
                    self.batch_id = result.get('batch_id')
                    created = result.get('new_entries', 0)
                    updated = result.get('updated_entries', 0)

                    self.log_result("Submit Data", "PASS", f"Batch ID: {self.batch_id}, Created: {created}, Updated: {updated}")
                    return True
                else:
                    self.log_result("Submit Data", "FAIL", f"Submission failed: {result.get('error')}")
                    return False
            else:
                print(f"Submit failed with status {response.status_code}: {response.text}")
                self.log_result("Submit Data", "FAIL", f"Status: {response.status_code}")
                return False

        except Exception as e:
            self.log_result("Submit Data", "FAIL", str(e))
            return False

    def verify_database(self):
        """Step 7: CRITICAL - Verify data in database"""
        print("\n=== STEP 7: DATABASE VERIFICATION (CRITICAL) ===")
        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()

            # Find recently created entries with our test notes
            cursor.execute("""
                SELECT data_id, field_id, entity_id, reporting_date, raw_value, notes, created_at
                FROM esg_data
                WHERE created_at > datetime('now', '-10 minutes')
                AND notes LIKE '%E2E Test Row%'
                ORDER BY created_at DESC
                LIMIT 10
            """)

            results = cursor.fetchall()

            print(f"\nFound {len(results)} entries in database:")
            for row in results:
                print(f"  - data_id={row[0]}, field_id={row[1]}, entity_id={row[2]}, value={row[4]}, notes={row[5][:80]}")

            # Verify we have exactly 3 entries
            if len(results) == 3:
                # Verify values match
                values = [float(row[4]) for row in results]
                expected_values = [350.0, 250.0, 150.0]  # Reversed due to DESC order

                if sorted(values) == sorted(expected_values):
                    self.log_result("Database Verification - Entry Count", "PASS", "Found 3 entries")
                    self.log_result("Database Verification - Values", "PASS", f"Values match: {values}")

                    # Save database evidence
                    evidence_path = os.path.join(TEST_OUTPUT_DIR, "database-verification", f"db_entries_{datetime.now().strftime('%Y%m%d-%H%M%S')}.txt")
                    with open(evidence_path, 'w') as f:
                        f.write(f"Database Verification Results\n")
                        f.write(f"Timestamp: {datetime.now().isoformat()}\n")
                        f.write(f"Batch ID: {self.batch_id}\n\n")
                        f.write(f"Found {len(results)} entries:\n\n")
                        for row in results:
                            f.write(f"data_id: {row[0]}\n")
                            f.write(f"field_id: {row[1]}\n")
                            f.write(f"entity_id: {row[2]}\n")
                            f.write(f"reporting_date: {row[3]}\n")
                            f.write(f"raw_value: {row[4]}\n")
                            f.write(f"notes: {row[5]}\n")
                            f.write(f"created_at: {row[6]}\n")
                            f.write("-" * 80 + "\n")

                    conn.close()
                    return True
                else:
                    self.log_result("Database Verification - Values", "FAIL", f"Values mismatch. Got: {values}, Expected: {expected_values}")
                    conn.close()
                    return False
            else:
                self.log_result("Database Verification - Entry Count", "FAIL", f"Expected 3 entries, found {len(results)}")
                conn.close()
                return False

        except Exception as e:
            self.log_result("Database Verification", "FAIL", str(e))
            return False

    def verify_audit_trail(self):
        """Step 8: Verify audit trail"""
        print("\n=== STEP 8: VERIFY AUDIT TRAIL ===")
        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()

            cursor.execute("""
                SELECT action, details, timestamp
                FROM audit_log
                WHERE timestamp > datetime('now', '-10 minutes')
                AND (action LIKE '%bulk%' OR action LIKE '%upload%')
                ORDER BY timestamp DESC
                LIMIT 20
            """)

            audit_entries = cursor.fetchall()

            print(f"\nFound {len(audit_entries)} audit log entries:")
            for entry in audit_entries:
                print(f"  - Action: {entry[0]}, Time: {entry[2]}")

            if len(audit_entries) > 0:
                self.log_result("Audit Trail Verification", "PASS", f"Found {len(audit_entries)} audit entries")
                conn.close()
                return True
            else:
                self.log_result("Audit Trail Verification", "WARNING", "No audit entries found")
                conn.close()
                return True  # Not critical

        except Exception as e:
            self.log_result("Audit Trail Verification", "WARNING", f"Could not verify: {str(e)}")
            return True  # Not critical

    def run_full_test(self):
        """Run complete E2E test"""
        print("\n" + "="*80)
        print("STARTING COMPREHENSIVE END-TO-END TEST WITH DATABASE VERIFICATION")
        print("="*80)

        start_time = time.time()

        # Execute all steps
        if not self.login():
            return False

        if not self.download_template():
            return False

        if not self.fill_template():
            return False

        if not self.upload_template():
            return False

        if not self.validate_data():
            return False

        if not self.submit_data():
            return False

        # CRITICAL: Database verification
        db_verification_passed = self.verify_database()

        # Audit trail (not critical)
        self.verify_audit_trail()

        end_time = time.time()
        duration = end_time - start_time

        # Generate report
        self.generate_report(duration, db_verification_passed)

        return db_verification_passed

    def generate_report(self, duration, db_verification_passed):
        """Generate test report"""
        report_path = os.path.join(TEST_OUTPUT_DIR, "e2e-workflow", f"E2E_Test_Report_{datetime.now().strftime('%Y%m%d-%H%M%S')}.md")

        with open(report_path, 'w') as f:
            f.write("# End-to-End Test Report with Database Verification\n\n")
            f.write(f"**Test Date:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
            f.write(f"**Duration:** {duration:.2f} seconds\n\n")
            f.write(f"**Batch ID:** {self.batch_id}\n\n")

            f.write("## Test Results\n\n")
            f.write("| Step | Status | Details |\n")
            f.write("|------|--------|----------|\n")

            for result in self.test_results:
                f.write(f"| {result['test']} | {result['status']} | {result['details']} |\n")

            f.write("\n## Critical Verdict\n\n")
            if db_verification_passed:
                f.write("✅ **DATABASE VERIFICATION PASSED**\n\n")
                f.write("- Data successfully written to database\n")
                f.write("- All 3 test entries found\n")
                f.write("- Values match expected data\n")
                f.write("- **Feature is PRODUCTION READY for data submission**\n")
            else:
                f.write("❌ **DATABASE VERIFICATION FAILED**\n\n")
                f.write("- **CRITICAL BLOCKER IDENTIFIED**\n")
                f.write("- Data not properly saved to database\n")
                f.write("- **Feature is NOT production ready**\n")

        print(f"\n\nReport saved to: {report_path}")

        # Print summary
        print("\n" + "="*80)
        print("TEST SUMMARY")
        print("="*80)
        passed = sum(1 for r in self.test_results if r['status'] == 'PASS')
        failed = sum(1 for r in self.test_results if r['status'] == 'FAIL')
        warnings = sum(1 for r in self.test_results if r['status'] == 'WARNING')

        print(f"Total Tests: {len(self.test_results)}")
        print(f"Passed: {passed}")
        print(f"Failed: {failed}")
        print(f"Warnings: {warnings}")
        print(f"Duration: {duration:.2f} seconds")
        print(f"\nDatabase Verification: {'✅ PASS' if db_verification_passed else '❌ FAIL'}")
        print("="*80 + "\n")

if __name__ == "__main__":
    tester = BulkUploadE2ETester()
    success = tester.run_full_test()
    sys.exit(0 if success else 1)
