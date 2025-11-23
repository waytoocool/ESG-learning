#!/usr/bin/env python3
"""
Quick End-to-End Test for Enhancement #4: Bulk Excel Upload
Tests complete workflow: Download → Fill → Upload → Validate → Submit → Verify DB
"""

import requests
import openpyxl
import sqlite3
import time
from datetime import datetime

# Configuration
BASE_URL = "http://test-company-alpha.127-0-0-1.nip.io:8000"
USERNAME = "bob@alpha.com"
PASSWORD = "user123"
DB_PATH = "instance/esg_data.db"

print("="*70)
print("QUICK END-TO-END TEST - Enhancement #4: Bulk Excel Upload")
print("="*70)

# Step 1: Login
print("\n[1/8] Logging in...")
session = requests.Session()
login_response = session.post(
    f"{BASE_URL}/login",
    data={"email": USERNAME, "password": PASSWORD},
    allow_redirects=True
)

if login_response.status_code not in [200, 302]:
    print(f"❌ Login failed: {login_response.status_code}")
    print(f"Response: {login_response.text[:200]}")
    exit(1)
print("✅ Logged in successfully")

# Step 2: Download Template
print("\n[2/8] Downloading template...")
template_response = session.post(
    f"{BASE_URL}/api/user/v2/bulk-upload/template",
    json={"filter": "pending"}
)

if template_response.status_code != 200:
    print(f"❌ Template download failed: {template_response.status_code}")
    print(f"Response: {template_response.text}")
    exit(1)

template_filename = f"quick_test_template_{int(time.time())}.xlsx"
with open(template_filename, 'wb') as f:
    f.write(template_response.content)
print(f"✅ Template downloaded: {template_filename}")

# Step 3: Fill Template
print("\n[3/8] Filling template with test data...")
wb = openpyxl.load_workbook(template_filename)
ws = wb['Data Entry']

timestamp = str(int(time.time()))
test_identifier = f"QUICK-E2E-TEST-{timestamp}"

# Fill 3 rows with test data
for idx in range(2, 5):  # Rows 2-4
    ws.cell(row=idx, column=4).value = 100 + idx  # Value column
    ws.cell(row=idx, column=6).value = f'{test_identifier}-ROW-{idx}'  # Notes column

filled_filename = f"quick_test_filled_{timestamp}.xlsx"
wb.save(filled_filename)
print(f"✅ Template filled: {filled_filename}")

# Step 4: Upload File
print("\n[4/8] Uploading filled template...")
with open(filled_filename, 'rb') as f:
    upload_response = session.post(
        f"{BASE_URL}/api/user/v2/bulk-upload/upload",
        files={"file": (filled_filename, f, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
    )

if upload_response.status_code != 200:
    print(f"❌ Upload failed: {upload_response.status_code}")
    print(f"Response: {upload_response.text}")
    exit(1)

upload_data = upload_response.json()
if not upload_data.get('success'):
    print(f"❌ Upload failed: {upload_data.get('error')}")
    exit(1)

upload_id = upload_data['upload_id']
print(f"✅ File uploaded successfully: upload_id={upload_id}")

# Step 5: Validate Data
print("\n[5/8] Validating uploaded data...")
validate_response = session.post(
    f"{BASE_URL}/api/user/v2/bulk-upload/validate",
    json={"upload_id": upload_id}
)

if validate_response.status_code != 200:
    print(f"❌ Validation failed: {validate_response.status_code}")
    print(f"Response: {validate_response.text}")
    exit(1)

validate_data = validate_response.json()
if not validate_data.get('success'):
    print(f"❌ Validation failed: {validate_data.get('error')}")
    exit(1)

print(f"✅ Validation passed:")
print(f"   - Valid rows: {validate_data.get('valid_count', 0)}")
print(f"   - Invalid rows: {validate_data.get('invalid_count', 0)}")

# Step 6: Submit Data
print("\n[6/8] Submitting validated data...")
submit_response = session.post(
    f"{BASE_URL}/api/user/v2/bulk-upload/submit",
    data={"upload_id": upload_id}
)

if submit_response.status_code != 200:
    print(f"❌ Submission failed: {submit_response.status_code}")
    print(f"Response: {submit_response.text}")
    exit(1)

submit_data = submit_response.json()
if not submit_data.get('success'):
    print(f"❌ Submission failed: {submit_data.get('error')}")
    exit(1)

print(f"✅ Data submitted successfully:")
print(f"   - Batch ID: {submit_data.get('batch_id')}")
print(f"   - New entries: {submit_data.get('new_entries', 0)}")
print(f"   - Updated entries: {submit_data.get('updated_entries', 0)}")

# Step 7: Verify Database
print("\n[7/8] Verifying data in database...")
conn = sqlite3.connect(DB_PATH)
cursor = conn.cursor()

cursor.execute("""
    SELECT data_id, raw_value, notes, created_at
    FROM esg_data
    WHERE notes LIKE ?
    ORDER BY created_at DESC
    LIMIT 10
""", (f'%{test_identifier}%',))

results = cursor.fetchall()
conn.close()

if len(results) == 0:
    print(f"❌ No entries found in database (expected 3)")
    exit(1)

print(f"✅ Database verification:")
print(f"   - Entries found: {len(results)}")
for row in results:
    print(f"   - data_id={row[0]}, value={row[1]}, notes={row[2][:50]}...")

# Step 8: Final Verification
print("\n[8/8] Final verification...")
if len(results) == 3:
    print("\n" + "="*70)
    print("✅ ✅ ✅  ALL TESTS PASSED - FEATURE WORKING END-TO-END  ✅ ✅ ✅")
    print("="*70)
    print(f"\n✓ Template downloaded successfully")
    print(f"✓ Template filled with 3 rows")
    print(f"✓ File uploaded and parsed")
    print(f"✓ Data validated successfully")
    print(f"✓ Data submitted to database")
    print(f"✓ 3 entries created in database")
    print(f"\nFeature is PRODUCTION READY!")
else:
    print(f"\n⚠️  Partial success: Expected 3 entries, found {len(results)}")
    exit(1)
